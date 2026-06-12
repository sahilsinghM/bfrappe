"""Import 185 commission slabs from the Top Slab Excel workbook.

Pure parsing/resolution logic is in parse_slab_row() and resolve_lender_entity()
— testable without Frappe. import_slabs() wires them to Frappe DocTypes.

Usage:
  bench --site <site> execute basic_spine.spine.make_slabs.import_slabs \
    --kwargs "{'path':'sites/prototype_data/Top Slab (2).xlsx'}"
"""

from rapidfuzz import process as fz_process
import openpyxl


_FUZZY_THRESHOLD = 70  # minimum RapidFuzz score to accept a match


def parse_slab_row(row: dict, sheet: str) -> dict | None:
    """Convert a raw Excel row dict into a Commission Slab field dict.

    Returns None for rows with no bank name (blank/header rows).
    """
    bank_name = (row.get("Bank name") or row.get("Bank Name") or "").strip()
    if not bank_name:
        return None

    top_slab_raw = row.get("Top-Slab") or row.get("Top Slab") or ""
    try:
        top_rate = float(top_slab_raw)
        rate_formula = ""
    except (TypeError, ValueError):
        top_rate = None
        rate_formula = str(top_slab_raw).strip()

    payout_type = "Gross" if (sheet or "").strip().lower() == "gross" else "Net"

    processing_raw = (row.get("Central/Manual") or "").strip().capitalize()
    if processing_raw not in ("Central", "Manual"):
        processing_raw = "Central"

    return {
        "bank_name": bank_name,
        "product": (row.get("Product") or "").strip(),
        "basic_dsa_code": (row.get("Bank Code") or "").strip(),
        "top_rate": top_rate,
        "rate_formula": rate_formula,
        "condition": (row.get("Condition") or "").strip(),
        "processing_mode": processing_raw,
        "payout_type": payout_type,
    }


def resolve_lender_entity(
    bank_name: str,
    known_entities: dict[str, str],
    aliases: dict[str, str],
    threshold: int = _FUZZY_THRESHOLD,
) -> str | None:
    """Resolve a raw bank name from the Excel to a Lender Entity name.

    Resolution order:
      1. Exact match in known_entities
      2. Alias lookup
      3. RapidFuzz best match in known_entities above threshold
      4. None (caller should create a skeleton record)
    """
    if bank_name in known_entities:
        return known_entities[bank_name]
    if bank_name in aliases:
        return aliases[bank_name]

    best = fz_process.extractOne(bank_name, list(known_entities.keys()))
    if best and best[1] >= threshold:
        return known_entities[best[0]]
    return None


def _read_sheet(ws) -> list[dict]:
    """Convert an openpyxl worksheet into a list of dicts using the header row."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    return [
        {headers[i]: cell for i, cell in enumerate(row)}
        for row in rows[1:]
        if any(cell is not None and str(cell).strip() for cell in row)
    ]


def import_slabs(path: str) -> dict:
    import frappe
    """Read the Top Slab Excel and upsert Commission Slab records.

    Returns summary: {"imported": int, "skipped": int, "unresolved": list}.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    known_entities = {
        e: e for e in frappe.get_all("Lender Entity", pluck="entity_name")
    }
    raw_aliases = frappe.get_all(
        "Lender Alias", fields=["raw_label", "lender_entity"]
    )
    aliases = {a.raw_label: a.lender_entity for a in raw_aliases}

    imported = skipped = 0
    unresolved: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = _read_sheet(ws)
        for row in rows:
            parsed = parse_slab_row(row, sheet=sheet_name)
            if parsed is None:
                continue

            entity = resolve_lender_entity(parsed["bank_name"], known_entities, aliases)
            if entity is None:
                if parsed["bank_name"] not in unresolved:
                    unresolved.append(parsed["bank_name"])
                    _create_skeleton(parsed["bank_name"])
                skipped += 1
                continue

            # Idempotency key: entity + product + payout_type
            exists = frappe.db.exists("Commission Slab", {
                "lender_entity": entity,
                "product": parsed["product"] or None,
                "payout_type": parsed["payout_type"],
                "basic_dsa_code": parsed["basic_dsa_code"] or None,
            })
            if exists:
                skipped += 1
                continue

            frappe.get_doc({
                "doctype": "Commission Slab",
                "lender_entity": entity,
                "product": parsed["product"] or None,
                "basic_dsa_code": parsed["basic_dsa_code"],
                "top_rate": parsed["top_rate"],
                "rate_formula": parsed["rate_formula"],
                "condition": parsed["condition"],
                "processing_mode": parsed["processing_mode"],
                "payout_type": parsed["payout_type"],
            }).insert(ignore_permissions=True)
            imported += 1

    frappe.db.commit()
    if unresolved:
        frappe.log_error(
            title="Commission Slab import — unresolved lenders",
            message="\n".join(unresolved),
        )
    return {"imported": imported, "skipped": skipped, "unresolved": unresolved}


def _create_skeleton(bank_name: str) -> None:
    import frappe
    """Create inactive Lender + Lender Entity stubs for unresolved names."""
    lender_name = bank_name[:140]
    if not frappe.db.exists("Lender", lender_name):
        frappe.get_doc({
            "doctype": "Lender",
            "lender_name": lender_name,
            "status": "Inactive",
        }).insert(ignore_permissions=True)
    if not frappe.db.exists("Lender Entity", bank_name[:140]):
        frappe.get_doc({
            "doctype": "Lender Entity",
            "entity_name": bank_name[:140],
            "lender": lender_name,
            "is_active": 0,
        }).insert(ignore_permissions=True)
