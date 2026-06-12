"""Synthetic prototype_data generator.

Real BASIC exports contain PII and stay out of git; these synthetic files
reproduce every documented edge case instead (zero-width Unicode in ids,
backslashed LANs, AU L/HFT- prefixes, LAN/AppId column swaps, drip-feed
partial MIS files, a line with no LAN, claimed-post-MIS reporting).

bench --site <site> execute basic_spine.spine.make_fixtures.make_all \
    --kwargs "{'out_dir':'prototype_data'}"

Also runnable without Frappe:  python -m basic_spine.spine.make_fixtures
"""

import os
import random

from openpyxl import Workbook

ZWSP = "\u200b"  # zero-width space

REPORTED_HEADERS = [
    "Team Name", "Basic id", "Disb id", "Lender App id", "Lender Disb id",
    "Customer Name", "pincode", "District", "Lender", "application_date",
    "Login Date", "Login Amt.", "Sanc Date", "Sanc Amt.", "Disb Date",
    "Disb. Created on", "Disb. Amt.", "Disb. Status", "Disb. Link",
    "Agent", "Agent RM", "Ful. RM",
]

FIRST = ["Ramesh", "Sunita", "Amit", "Priya", "Vikas", "Neha", "Rajesh", "Pooja",
         "Sandeep", "Kavita", "Manoj", "Anita", "Deepak", "Rekha", "Suresh",
         "Geeta", "Arun", "Lakshmi", "Vinod", "Meena"]
LAST = ["Kumar", "Sharma", "Verma", "Singh", "Gupta", "Yadav", "Patel", "Reddy",
        "Joshi", "Mehta", "Agarwal", "Mishra", "Chauhan", "Nair", "Das"]
DISTRICTS = ["Gurgaon", "Faridabad", "Jaipur", "Lucknow", "Indore", "Pune",
             "Nagpur", "Surat", "Bhopal", "Patna"]
TEAMS = ["North-1", "North-2", "West-1", "Central-1"]


def _name(rng):
    return f"{rng.choice(FIRST)} {rng.choice(LAST)}"


def _fmt_date(d, style):
    """Mix the two real-world date string formats: 01-Apr-2026 and 07-Apr-26."""
    return d.strftime("%d-%b-%Y") if style == 0 else d.strftime("%d-%b-%y")


class Case:
    """One reported disbursement and (optionally) its MIS counterpart."""

    seq = 0

    def __init__(self, rng, lender_label, scenario, tranche=1, case_id=None):
        Case.seq += 1
        self.scenario = scenario
        self.lender_label = lender_label
        self.case_id = case_id or f"B{Case.seq:05d}X"
        self.disb_id = f"{self.case_id}_D{tranche}"
        self.customer = _name(rng)
        self.amount = float(rng.randrange(8, 160) * 50000)
        import datetime
        if scenario == "overdue":
            self.disb_date = datetime.date(2026, rng.choice([2, 3]), rng.randrange(1, 28))
        else:
            self.disb_date = datetime.date(2026, 4, rng.randrange(1, 14))
        self.created_on = self.disb_date + datetime.timedelta(days=rng.randrange(0, 3))
        if scenario == "claimed_post":
            self.created_on = datetime.date(2026, 4, rng.randrange(25, 30))
        self.status = "VerifiedByBasic"
        self.app_id = None
        self.lan = None
        self.mis_lan = None
        self.mis_app = None


def _chola_cases(rng):
    """Chola: LAN HE01GUF…, app ids 9-digit. Builds the scenario mix."""
    cases = []

    def lan(i):
        return f"HE01GUF{160600 + i:011d}"

    def app(i):
        return str(300965000 + i)

    i = 0
    # A: LAN exact — some raw values corrupted on the reported side (rescued)
    for k in range(10):
        i += 1
        c = Case(rng, "Chola", "A")
        c.app_id, c.lan, c.mis_lan, c.mis_app = app(i), lan(i), lan(i), app(i)
        if k < 3:  # zero-width corruption on reported side only -> norm rescues
            c.lan = lan(i) + ZWSP
            c.scenario = "A_rescued"
        elif k == 3:  # backslash corruption
            c.lan = lan(i)[:4] + "\\" + lan(i)[4:7] + "\\" + lan(i)[7:]
            c.scenario = "A_rescued"
        cases.append(c)

    # A-: MIS has no LAN, joins on app id (one written with hyphens -> rescued)
    for k in range(3):
        i += 1
        c = Case(rng, "Chola", "A-")
        c.app_id, c.lan = app(i), lan(i)
        c.mis_lan, c.mis_app = None, (f"{app(i)[:3]}-{app(i)[3:6]}-{app(i)[6:]}" if k == 0 else app(i))
        if k == 0:
            c.scenario = "A-_rescued"
        cases.append(c)

    # cross-swap: MIS put the app id in its LAN column
    i += 1
    c = Case(rng, "Chola", "cross")
    c.app_id, c.lan = app(i), lan(i)
    c.mis_lan, c.mis_app = app(i), None
    cases.append(c)

    # B: ids useless; amount/date/name triangulation
    for _ in range(2):
        i += 1
        c = Case(rng, "Chola", "B")
        c.app_id, c.lan = app(i), lan(i)
        c.mis_lan, c.mis_app = "NA", None
        cases.append(c)

    # C: name + amount(±5%) only — no date, ids missing
    for _ in range(2):
        i += 1
        c = Case(rng, "Chola", "C")
        c.app_id, c.lan = app(i), lan(i)
        c.mis_lan, c.mis_app = None, None
        cases.append(c)

    # ClaimedPostMIS: business reported after the MIS arrived
    for _ in range(2):
        i += 1
        c = Case(rng, "Chola", "claimed_post")
        c.app_id, c.lan, c.mis_lan, c.mis_app = app(i), lan(i), lan(i), app(i)
        cases.append(c)

    # Multi-tranche case: same basic_case_id, two disbursements, two MIS lines
    i += 1
    t1 = Case(rng, "Chola", "A", tranche=1, case_id="B00TRN")
    t1.app_id = app(i)
    t1.lan, t1.mis_lan, t1.mis_app = lan(i), lan(i), app(i)
    i += 1
    t2 = Case(rng, "Chola", "A", tranche=2, case_id="B00TRN")
    t2.app_id = t1.app_id  # app id repeats across tranches (ICICI-style reality)
    t2.lan, t2.mis_lan, t2.mis_app = lan(i), lan(i), t1.app_id
    cases += [t1, t2]

    # Reported but never in MIS -> AwaitingMIS / Overdue book
    for k in range(5):
        i += 1
        c = Case(rng, "Chola", "overdue" if k < 3 else "awaiting")
        c.app_id, c.lan = app(i), lan(i)
        cases.append(c)

    return cases


def _au_cases(rng):
    """AU: 16-digit LANs that MIS writes with a leading L; HFT- app prefixes."""
    cases = []

    def lan(i):
        return str(9001061756938000 + i)

    def app(i):
        return str(1110200 + i)

    i = 0
    for k in range(8):  # A via LAN; MIS adds the L prefix (rescued via transform)
        i += 1
        c = Case(rng, "AU SFB", "A_rescued" if k < 5 else "A")
        c.app_id, c.lan = f"HFT-{app(i)}", lan(i)
        c.mis_lan = ("L" + lan(i)) if k < 5 else lan(i)
        c.mis_app = f"HFT-{app(i)}"
        cases.append(c)

    for _ in range(2):  # A-: no LAN in MIS, app id joins (HFT- stripped by norm)
        i += 1
        c = Case(rng, "AU SFB", "A-")
        c.app_id, c.lan = f"HFT-{app(i)}", lan(i)
        c.mis_lan, c.mis_app = None, str(app(i))  # MIS drops the HFT- prefix
        cases.append(c)

    i += 1  # the real April case: one line with no LAN at all
    c = Case(rng, "AU SFB", "A-")
    c.app_id, c.lan = f"HFT-{app(i)}", lan(i)
    c.mis_lan, c.mis_app = None, f"HFT-{app(i)}"
    cases.append(c)

    for k in range(4):  # never in MIS
        i += 1
        c = Case(rng, "AU SFB", "overdue" if k < 2 else "awaiting")
        c.app_id, c.lan = f"HFT-{app(i)}", lan(i)
        cases.append(c)

    return cases


def _other_lender_rows(rng):
    """Reported-only rows for the big lenders (no MIS files in the prototype)."""
    specs = [
        ("HDFC Sales", lambda i: str(704335800 + i), None, 8),
        ("ICICI Bank", lambda i: str(77212306100 + i), None, 7),
        ("Tata Capital Finance", lambda i: f"TCHHL{453000100415400 + i:016d}", None, 5),
        ("Aditya Birla Housing Finance",
         lambda i: (f"LNDELLAP-0725030{7400 + i}" if i % 2 else str(3000142000 + i)), None, 5),
        ("PNB Housing Finance", lambda i: str(10409800 + i), None, 4),
        ("Bajaj Housing Finance", lambda i: f"H400HHL{1553400 + i}", None, 4),
        ("ISFC (India Shelter)", lambda i: str(10409900 + i), None, 3),
        ("Mahindra Rural (New)", lambda i: str(880077000 + i), None, 3),  # unresolvable label
    ]
    cases = []
    for label, app_fn, _, count in specs:
        for i in range(count):
            scen = "overdue" if (label != "Mahindra Rural (New)" and i == 0) else "awaiting"
            c = Case(rng, label, scen)
            c.app_id = app_fn(i)
            c.lan = None
            if rng.random() < 0.22:
                c.status = rng.choice(["Cancelled", "RejectedByBasic"])
            cases.append(c)
    return cases


def _write_reported(cases, path, rng):
    wb = Workbook()
    ws = wb.active
    ws.title = "Case Pipeline"
    # Real export has a 2-row banner; data header is the 3rd row.
    ws.append(["BASIC - Case Pipeline Export"] + [None] * (len(REPORTED_HEADERS) - 1))
    ws.append(["Generated: 12-Jun-2026 09:14 | Filters: FY26"] + [None] * (len(REPORTED_HEADERS) - 1))
    ws.append(REPORTED_HEADERS)
    import datetime
    for c in cases:
        style = rng.randrange(2)
        app_d = c.disb_date - datetime.timedelta(days=rng.randrange(20, 60))
        login_d = app_d + datetime.timedelta(days=rng.randrange(1, 5))
        sanc_d = login_d + datetime.timedelta(days=rng.randrange(5, 15))
        ws.append([
            rng.choice(TEAMS), c.case_id, c.disb_id, c.app_id, c.lan,
            c.customer, str(rng.randrange(110001, 855118)), rng.choice(DISTRICTS),
            c.lender_label, _fmt_date(app_d, style), _fmt_date(login_d, style),
            c.amount * 1.1, _fmt_date(sanc_d, style), c.amount * 1.05,
            _fmt_date(c.disb_date, style), _fmt_date(c.created_on, style),
            c.amount, c.status, f"https://console.basic.example/case/{c.case_id}",
            f"AG{rng.randrange(1000, 9999)}", f"RM{rng.randrange(100, 999)}",
            f"FRM{rng.randrange(100, 999)}",
        ])
    wb.save(path)
    return len(cases)


def _chola_mis_row(c, rng, garble_name=False):
    import datetime
    name = c.customer.upper()
    if garble_name:
        # light corruption: an extra initial, as ops typists do; keeps
        # token_sort_ratio above the fuzzy thresholds while raw strings differ
        name = f"{name} {rng.choice('RSKM')}"
    amt = c.amount
    date = c.disb_date
    if c.scenario == "B":
        amt = round(c.amount * rng.uniform(0.995, 1.005), 0)  # within 1%
        date = c.disb_date + datetime.timedelta(days=rng.choice([-2, 2, 3]))
    if c.scenario == "C":
        amt = round(c.amount * rng.uniform(1.02, 1.04), 0)  # within 5%, beyond 1%
        date = None
    payin = round(amt * 0.011, 2)
    return [
        c.mis_lan, c.mis_app, name, "HE01", "GURGAON FIVE HE", "Haryana",
        round(amt * 1.05, 0), amt,
        date.strftime("%d/%m/%Y") if date else None,
        payin, 1.1, None,
    ]


def _write_chola_mis(cases, path, rng):
    wb = Workbook()
    ws = wb.active
    ws.title = "MISReportLines"
    ws.append(["Loan Account No", "Application No", "Customer Name", "Product",
               "Branch Name", "State", "Sanction Amount", "Disbursal Amount",
               "Disbursal Date", "Payout Amount", "Payout %", "Remarks"])
    n = 0
    for c in cases:
        if c.scenario in ("awaiting", "overdue"):
            continue  # never reaches the MIS — that's the leakage book
        ws.append(_chola_mis_row(c, rng, garble_name=(c.scenario in ("B", "C"))))
        n += 1
    # leakage: lines BASIC never reported (unknown to the console)
    for i in range(3):
        ws.append([f"HE01JPR{990000 + i:011d}", str(300999900 + i),
                   _name(rng).upper(), "HE01", "JAIPUR TWO HE", "Rajasthan",
                   2100000 + i * 100000, 2000000 + i * 100000, "10/04/2026",
                   22000.0, 1.1, "No console record"])
        n += 1
    ws2 = wb.create_sheet("Instructions")
    ws2.append(["Do not edit. System generated payout MIS for Apr-26 cycle."])
    wb.save(path)
    return n


def _au_row(c):
    return [c.mis_lan, c.mis_app, c.customer.upper(), "HL", "GURGAON",
            c.amount, c.disb_date.strftime("%d-%m-%Y"), round(c.amount * 0.009, 2)]


def _write_au_mis(cases, path1, path2, rng):
    headers = ["LAN", "APPL NO", "CUSTOMER NAME", "PRODUCT", "BRANCH",
               "DISB AMT", "DISB DATE", "NET PAYOUT"]
    mis_cases = [c for c in cases if c.scenario not in ("awaiting", "overdue")]
    part1 = mis_cases[:7]
    # Drip feed: part 2 re-sends 3 of part 1's lines plus the remaining new ones
    part2 = mis_cases[4:]

    for path, rows in ((path1, part1), (path2, part2)):
        wb = Workbook()
        ws = wb.active
        ws.title = "Payout"
        ws.append(headers)
        for c in rows:
            ws.append(_au_row(c))
        wb.save(path)
    return len(part1), len(part2)


def make_all(out_dir="prototype_data"):
    rng = random.Random(7)
    Case.seq = 0
    os.makedirs(out_dir, exist_ok=True)

    chola = _chola_cases(rng)
    au = _au_cases(rng)
    others = _other_lender_rows(rng)

    n_rep = _write_reported(chola + au + others, os.path.join(out_dir, "reported.xlsx"), rng)
    n_chola = _write_chola_mis(chola, os.path.join(out_dir, "mis_chola.xlsx"), rng)
    n_au1, n_au2 = _write_au_mis(au, os.path.join(out_dir, "mis_au_part1.xlsx"),
                                 os.path.join(out_dir, "mis_au_part2.xlsx"), rng)

    print(f"[make_fixtures] {out_dir}/reported.xlsx: {n_rep} rows")
    print(f"[make_fixtures] {out_dir}/mis_chola.xlsx: {n_chola} lines (sheet MISReportLines)")
    print(f"[make_fixtures] {out_dir}/mis_au_part1.xlsx: {n_au1} lines, "
          f"mis_au_part2.xlsx: {n_au2} lines (drip-feed overlap)")
    return out_dir


if __name__ == "__main__":
    make_all()
